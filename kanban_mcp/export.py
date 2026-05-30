#!/usr/bin/env python3
"""
Kanban Export Module
Provides multi-format export functionality
(JSON, YAML, Markdown, XLSX) for kanban data.
"""

from io import BytesIO
import json
from datetime import datetime
from typing import Dict, List, Any, Optional

STATUS_COLUMN_ORDER = [
    'חדשים',
    'ראיון טלפוני',
    'ראיון מקצועי',
    'רותם',
    'שכר',
    'סיווג',
    'אין התאמה',
]

# XLSX is optional - gracefully handle missing import
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# YAML is optional - gracefully handle missing import
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


class ExportBuilder:
    """Gathers kanban data with filtering options for export."""

    def __init__(self, db, project_id: str):
        """Initialize exporter with database connection and project.

        Args:
            db: KanbanDB instance
            project_id: Project ID to export
        """
        self.db = db
        self.project_id = project_id
        self.project = db.get_project_by_id(project_id)

    def build_export_data(
        self,
        item_ids: List[int] = None,
        item_type: str = None,
        statuses: List[str] = None,
        include_tags: bool = True,
        include_relationships: bool = False,
        include_metrics: bool = False,
        include_updates: bool = False,
        include_epic_progress: bool = False,
        include_decisions: bool = False,
        limit: int = None
    ) -> Dict[str, Any]:
        """Build complete export data structure.

        Args:
            item_ids: Optional list of specific item IDs to include
            statuses: Filter by multiple statuses
                (חדשים, ראיון טלפוני, ראיון מקצועי, רותם, שכר, סיווג, אין התאמה)
            include_tags: Include tag data for each item
            include_relationships: Include relationship data
            include_metrics: Include calculated metrics
                (lead_time, cycle_time, etc.)
            include_updates: Include project updates
            include_epic_progress: Include epic progress stats
            include_decisions: Include decision history for items
            limit: Maximum items to export

        Returns:
            Complete export data dictionary
        """
        # Build metadata
        metadata = {
            "project_name": (
                self.project['name'] if self.project
                else "Unknown"
            ),
            "project_path": (
                self.project['directory_path']
                if self.project else None
            ),
            "exported_at": datetime.now().isoformat(),
            "filters": {
                "item_ids": item_ids,
                "item_type": item_type,
                "statuses": statuses,
            },
            "include_options": {
                "tags": include_tags,
                "relationships": include_relationships,
                "metrics": include_metrics,
                "updates": include_updates,
                "epic_progress": include_epic_progress,
                "decisions": include_decisions
            }
        }

        # Get items based on filters
        if item_ids:
            items = [self.db.get_item(item_id) for item_id in item_ids]
            items = [i for i in items if i is not None]
        else:
            status_filter = statuses if statuses else None
            items = self.db.list_items(
                project_id=self.project_id,
                type_name='cv',
                status_names=status_filter,
                limit=limit
            )

        # Process items and add optional data
        processed_items = []
        for item in items:
            processed = self._process_item(
                item,
                include_tags=include_tags,
                include_relationships=include_relationships,
                include_metrics=include_metrics,
                include_epic_progress=include_epic_progress,
                include_decisions=include_decisions
            )
            processed_items.append(processed)

        # Build export data
        export_data = {
            "metadata": metadata,
            "items": processed_items,
            "summary": self._build_summary(processed_items)
        }

        # Include updates if requested
        if include_updates:
            updates = self.db.get_updates(self.project_id, limit=100)
            export_data["updates"] = self._serialize_updates(updates)

        return export_data

    def _process_item(
        self,
        item: Dict,
        include_tags: bool,
        include_relationships: bool,
        include_metrics: bool,
        include_epic_progress: bool,
        include_decisions: bool = False
    ) -> Dict[str, Any]:
        """Process a single item for export."""
        processed = {
            "id": item['id'],
            "title": item['title'],
            "description": item.get('description'),
            "type_name": item['type_name'],
            "status_name": item['status_name'],
            "priority": item['priority'],
            "complexity": item.get('complexity'),
            "parent_id": item.get('parent_id'),
            "created_at": self._serialize_datetime(item.get('created_at')),
            "closed_at": self._serialize_datetime(item.get('closed_at'))
        }

        if include_tags:
            tags = self.db.get_item_tags(item['id'])
            processed["tags"] = [
                {"id": t['id'], "name": t['name'], "color": t['color']}
                for t in tags
            ]

        if include_relationships:
            rels = self.db.get_item_relationships(item['id'])
            processed["relationships"] = {
                "outgoing": [
                    {
                        "type": r['relationship_type'],
                        "target_id": r['related_item_id'],
                        "target_title": r['related_item_title']
                    }
                    for r in rels.get('outgoing', [])
                ],
                "incoming": [
                    {
                        "type": r['relationship_type'],
                        "source_id": r['related_item_id'],
                        "source_title": r['related_item_title']
                    }
                    for r in rels.get('incoming', [])
                ]
            }

        if include_metrics:
            metrics = self.db.get_item_metrics(item['id'])
            if metrics:
                processed["metrics"] = {
                    "lead_time": metrics.get('lead_time'),
                    "cycle_time": metrics.get('cycle_time'),
                    "time_in_each_status": metrics.get(
                        'time_in_each_status', {}
                    ),
                    "revert_count": metrics.get('revert_count', 0),
                    "current_age": metrics.get('current_age')
                }

        if include_epic_progress and item['type_name'] == 'epic':
            progress = self.db.get_epic_progress(item['id'])
            processed["epic_progress"] = {
                "total": progress['total'],
                "completed": progress['completed'],
                "percent": progress['percent'],
                "incomplete_items": progress['incomplete_items']
            }

        if include_decisions:
            decisions = self.db.get_item_decisions(item['id'])
            processed["decisions"] = [
                {
                    "id": d['id'],
                    "choice": d['choice'],
                    "rejected_alternatives": d.get('rejected_alternatives'),
                    "rationale": d.get('rationale'),
                    "created_at": self._serialize_datetime(d.get('created_at'))
                }
                for d in decisions
            ]

        return processed

    def _build_summary(self, items: List[Dict]) -> Dict[str, Any]:
        """Build summary statistics from items."""
        by_type = {}
        by_status = {}

        for item in items:
            type_name = item['type_name']
            status_name = item['status_name']

            by_type[type_name] = by_type.get(type_name, 0) + 1
            by_status[status_name] = by_status.get(status_name, 0) + 1

        return {
            "total_items": len(items),
            "by_type": by_type,
            "by_status": by_status
        }

    def _serialize_updates(self, updates: List[Dict]) -> List[Dict]:
        """Serialize updates for export."""
        return [
            {
                "id": u['id'],
                "content": u['content'],
                "created_at": self._serialize_datetime(u.get('created_at')),
                "item_ids": u.get('item_ids', [])
            }
            for u in updates
        ]

    @staticmethod
    def _serialize_datetime(dt) -> Optional[str]:
        """Convert datetime to ISO string."""
        if dt is None:
            return None
        if isinstance(dt, datetime):
            return dt.isoformat()
        return str(dt)


def format_json(data: Dict[str, Any], indent: int = 2) -> str:
    """Format export data as JSON.

    Args:
        data: Export data dictionary
        indent: Indentation level (default 2)

    Returns:
        JSON string
    """
    return json.dumps(data, indent=indent, default=str)


def format_yaml(data: Dict[str, Any]) -> str:
    """Format export data as YAML.

    Args:
        data: Export data dictionary

    Returns:
        YAML string

    Raises:
        ImportError: If pyyaml is not installed
    """
    if not YAML_AVAILABLE:
        raise ImportError(
            "YAML export requires pyyaml. Install with: pip install pyyaml"
        )
    return yaml.dump(
        data, default_flow_style=False,
        allow_unicode=True, sort_keys=False
    )


def format_markdown(data: Dict[str, Any], detailed: bool = False) -> str:
    """Format export data as Markdown.

    Args:
        data: Export data dictionary
        detailed: If True, include full item details; if False, summary tables

    Returns:
        Markdown string
    """
    lines = []
    metadata = data.get("metadata", {})
    items = data.get("items", [])
    summary = data.get("summary", {})
    updates = data.get("updates", [])

    # Header
    project_name = metadata.get("project_name", "Kanban Export")
    lines.append(f"# Kanban Export: {project_name}")
    lines.append("")

    # Metadata
    exported_at = metadata.get("exported_at", "")
    if exported_at:
        # Format date nicely
        try:
            dt = datetime.fromisoformat(exported_at)
            exported_at = dt.strftime("%Y-%m-%d %H:%M")
        except (ValueError, TypeError):
            pass
    lines.append(f"**Exported:** {exported_at}")
    lines.append(f"**Total Items:** {summary.get('total_items', 0)}")
    lines.append("")

    # Filters applied
    filters = metadata.get("filters", {})
    active_filters = []
    if filters.get("statuses"):
        active_filters.append(
            f"statuses={','.join(filters['statuses'])}"
        )
    if filters.get("item_ids"):
        active_filters.append(f"ids={','.join(map(str, filters['item_ids']))}")
    if active_filters:
        lines.append(f"**Filters:** {', '.join(active_filters)}")
        lines.append("")

    # Summary section
    lines.append("## Summary")
    lines.append("")

    # By status
    by_status = summary.get("by_status", {})
    if by_status:
        lines.append("### By Status")
        for status in STATUS_COLUMN_ORDER:
            if status in by_status:
                lines.append(f"- **{status}:** {by_status[status]}")
        for status, count in sorted(by_status.items()):
            if status not in STATUS_COLUMN_ORDER:
                lines.append(f"- **{status}:** {count}")
        lines.append("")

    # Items section
    if detailed:
        lines.append("## Items (Detailed)")
        lines.append("")
        for item in items:
            lines.extend(_format_item_detailed(item))
            lines.append("")
    else:
        # Group items by status
        items_by_status = {}
        for item in items:
            status_name = item['status_name']
            if status_name not in items_by_status:
                items_by_status[status_name] = []
            items_by_status[status_name].append(item)

        lines.append("## Items by Status")
        lines.append("")

        for status_name in STATUS_COLUMN_ORDER:
            if status_name in items_by_status:
                lines.extend(_format_items_table(
                    status_name, items_by_status[status_name]
                ))
                lines.append("")

        # Any statuses not in the standard order
        for status_name in sorted(items_by_status.keys()):
            if status_name not in STATUS_COLUMN_ORDER:
                lines.extend(_format_items_table(
                    status_name, items_by_status[status_name]
                ))
                lines.append("")

    # Updates section
    if updates:
        lines.append("## Recent Updates")
        lines.append("")
        for update in updates[:20]:  # Limit to 20 updates
            created = update.get('created_at', '')
            if created:
                try:
                    dt = datetime.fromisoformat(created)
                    created = dt.strftime("%Y-%m-%d %H:%M")
                except (ValueError, TypeError):
                    pass
            content = update.get('content', '')
            item_ids = update.get('item_ids', [])
            if item_ids:
                items_str = ', '.join(f"#{i}" for i in item_ids)
                lines.append(f"- **{created}** [{items_str}]: {content}")
            else:
                lines.append(f"- **{created}**: {content}")
        lines.append("")

    return "\n".join(lines)


def _format_items_table(type_name: str, items: List[Dict]) -> List[str]:
    """Format items of a single status as a markdown table."""
    lines = []
    lines.append(f"### {type_name}")
    lines.append("")

    # Determine columns based on data
    has_tags = any(item.get('tags') for item in items)
    has_complexity = any(item.get('complexity') for item in items)

    # Header
    header = "| ID | Title | Status | Priority | Decisions |"
    separator = "|---|---|---|---|---|"
    if has_complexity:
        header += " Complexity |"
        separator += "---|"
    if has_tags:
        header += " Tags |"
        separator += "---|"

    lines.append(header)
    lines.append(separator)

    # Rows
    for item in items:
        title = (
            item['title'][:50] + "..."
            if len(item['title']) > 50
            else item['title']
        )
        # Escape pipe characters in title
        title = title.replace("|", "\\|")
        decisions = _summarize_decisions(item.get('decisions', []))
        row = (
            f"| #{item['id']} | {title} "
            f"| {item['status_name']} "
            f"| P{item['priority']} "
            f"| {decisions} |"
        )

        if has_complexity:
            complexity = (
                f"C{item['complexity']}"
                if item.get('complexity') else "-"
            )
            row += f" {complexity} |"

        if has_tags:
            tags = item.get('tags', [])
            tag_str = ', '.join(t['name'] for t in tags) if tags else "-"
            row += f" {tag_str} |"

        lines.append(row)

    return lines


def _format_item_detailed(item: Dict) -> List[str]:
    """Format a single item with full details."""
    lines = []

    lines.append(f"### #{item['id']} - {item['title']}")
    lines.append("")
    lines.append(f"- **Status:** {item['status_name']}")
    lines.append(f"- **Priority:** P{item['priority']}")

    if item.get('complexity'):
        lines.append(f"- **Complexity:** C{item['complexity']}")

    if item.get('parent_id'):
        lines.append(f"- **Parent:** #{item['parent_id']}")

    if item.get('created_at'):
        created = item['created_at']
        if isinstance(created, str):
            try:
                dt = datetime.fromisoformat(created)
                created = dt.strftime("%Y-%m-%d %H:%M")
            except (ValueError, TypeError):
                pass
        lines.append(f"- **Created:** {created}")

    if item.get('closed_at'):
        closed = item['closed_at']
        if isinstance(closed, str):
            try:
                dt = datetime.fromisoformat(closed)
                closed = dt.strftime("%Y-%m-%d %H:%M")
            except (ValueError, TypeError):
                pass
        lines.append(f"- **Closed:** {closed}")

    # Tags
    if item.get('tags'):
        tag_names = ', '.join(t['name'] for t in item['tags'])
        lines.append(f"- **Tags:** {tag_names}")

    # Description
    if item.get('description'):
        lines.append("")
        lines.append("**Description:**")
        lines.append("")
        lines.append(item['description'])

    # Relationships
    if item.get('relationships'):
        rels = item['relationships']
        if rels.get('outgoing') or rels.get('incoming'):
            lines.append("")
            lines.append("**Relationships:**")
            for rel in rels.get('outgoing', []):
                lines.append(
                    f"  - {rel['type']} "
                    f"→ #{rel['target_id']} "
                    f"({rel['target_title']})"
                )
            for rel in rels.get('incoming', []):
                lines.append(
                    f"  - #{rel['source_id']} "
                    f"({rel['source_title']}) "
                    f"→ {rel['type']}"
                )

    # Metrics
    if item.get('metrics'):
        metrics = item['metrics']
        lines.append("")
        lines.append("**Metrics:**")
        if metrics.get('lead_time') is not None:
            lines.append(f"  - Lead Time: {metrics['lead_time']} hours")
        if metrics.get('cycle_time') is not None:
            lines.append(f"  - Cycle Time: {metrics['cycle_time']} hours")
        if metrics.get('current_age') is not None:
            lines.append(f"  - Age: {metrics['current_age']} hours")
        if metrics.get('revert_count'):
            lines.append(f"  - Reverts: {metrics['revert_count']}")

    # Epic progress
    if item.get('epic_progress'):
        prog = item['epic_progress']
        lines.append("")
        lines.append(
            f"**Epic Progress:** "
            f"{prog['completed']}/{prog['total']} "
            f"({prog['percent']}%)"
        )
        if prog.get('incomplete_items'):
            incomplete = ', '.join(
                f"#{i}"
                for i in prog['incomplete_items'][:10]
            )
            if len(prog['incomplete_items']) > 10:
                incomplete += (
                    f" ... ({len(prog['incomplete_items']) - 10}"
                    " more)"
                )
            lines.append(f"  - Incomplete: {incomplete}")

    # Decision history
    if item.get('decisions'):
        lines.append("")
        lines.append("**Decisions:**")
        for d in item['decisions']:
            lines.append(f"  - **Chose:** {d['choice']}")
            if d.get('rejected_alternatives'):
                lines.append(f"    - Rejected: {d['rejected_alternatives']}")
            if d.get('rationale'):
                lines.append(f"    - Why: {d['rationale']}")

    return lines


def _summarize_decisions(decisions: List[Dict]) -> str:
    """Condense decision history for table views."""
    if not decisions:
        return "-"

    choices = [d.get('choice', '') for d in decisions if d.get('choice')]
    if not choices:
        return "-"

    summary = ' → '.join(choices[:3])
    if len(choices) > 3:
        summary += f" … (+{len(choices) - 3})"
    return summary


def export_to_format(
    data: Dict[str, Any],
    format: str = "json",
    detailed: bool = False
) -> str | bytes:
    """Export data to specified format.

    Args:
        data: Export data dictionary from ExportBuilder
        format: Output format ('json', 'yaml', 'markdown', 'xlsx')
        detailed: For markdown, include detailed item info

    Returns:
        Formatted string

    Raises:
        ValueError: If format is not supported
        ImportError: If yaml is requested but pyyaml not installed
    """
    format = format.lower()

    if format == "json":
        return format_json(data)
    elif format == "yaml":
        return format_yaml(data)
    elif format in ("markdown", "md"):
        return format_markdown(data, detailed=detailed)
    elif format == "xlsx":
        return format_xlsx(data)
    else:
        raise ValueError(
            f"Unsupported format: {format}. "
            "Use 'json', 'yaml', 'markdown', or 'xlsx'"
        )


def format_xlsx(data: Dict[str, Any]) -> bytes:
    """Format export data as an Excel workbook.

    Returns:
        XLSX binary content

    Raises:
        ImportError: If openpyxl is not installed
    """
    if not XLSX_AVAILABLE:
        raise ImportError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        )

    metadata = data.get("metadata", {})
    items = data.get("items", [])
    summary = data.get("summary", {})
    updates = data.get("updates", [])

    items_by_status = _group_items_by_status(items)
    ordered_statuses = _ordered_statuses(items_by_status)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "סיכום"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    section_fill = PatternFill("solid", fgColor="E5E7EB")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    project_name = metadata.get("project_name", "Kanban Export")
    summary_sheet["A1"] = f"דוח מועמדים - {project_name}"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet.merge_cells("A1:C1")

    summary_sheet["A2"] = "Exported At"
    summary_sheet["B2"] = metadata.get("exported_at", "")
    summary_sheet["A3"] = "Total Items"
    summary_sheet["B3"] = summary.get("total_items", 0)
    summary_sheet["A4"] = "By Status"
    summary_sheet["A4"].fill = section_fill

    by_status = summary.get("by_status", {})
    row = 5
    for status_name in STATUS_COLUMN_ORDER:
        if status_name in by_status:
            summary_sheet[f"A{row}"] = status_name
            summary_sheet[f"B{row}"] = by_status[status_name]
            row += 1
    for status_name in sorted(by_status.keys()):
        if status_name not in STATUS_COLUMN_ORDER:
            summary_sheet[f"A{row}"] = status_name
            summary_sheet[f"B{row}"] = by_status[status_name]
            row += 1

    status_sheet = workbook.create_sheet("סטטוסים")
    status_sheet.freeze_panes = "A3"
    status_sheet["A1"] = "מועמדים לפי סטטוס"
    status_sheet["A1"].font = Font(bold=True, size=14)
    status_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(ordered_statuses)))

    for col_idx, status_name in enumerate(ordered_statuses, start=1):
        header_cell = status_sheet.cell(row=2, column=col_idx, value=status_name)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_sheet.column_dimensions[header_cell.column_letter].width = 32

        count_cell = status_sheet.cell(
            row=3,
            column=col_idx,
            value=f"{len(items_by_status.get(status_name, []))} מועמדים"
        )
        count_cell.fill = section_fill
        count_cell.alignment = Alignment(horizontal="center")

        current_row = 4
        for item in items_by_status.get(status_name, []):
            card_cell = status_sheet.cell(
                row=current_row,
                column=col_idx,
                value=_format_status_card(item)
            )
            card_cell.alignment = wrap_top
            current_row += 1

    detail_sheet = workbook.create_sheet("פירוט")
    detail_sheet.freeze_panes = "A2"
    detail_columns = [
        ("ID", 10),
        ("Title", 30),
        ("Status", 18),
        ("Created At", 22),
        ("Age (Days)", 12),
        ("Decision History", 60),
    ]
    for col_idx, (title, width) in enumerate(detail_columns, start=1):
        cell = detail_sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        detail_sheet.column_dimensions[cell.column_letter].width = width

    for row_idx, item in enumerate(items, start=2):
        row_values = [
            item.get("id"),
            item.get("title"),
            item.get("status_name"),
            item.get("created_at"),
            _format_item_age_days(item.get("created_at")),
            _format_decision_history(item.get("decisions", [])),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = detail_sheet.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == len(row_values):
                cell.alignment = wrap_top

    if updates:
        updates_sheet = workbook.create_sheet("Updates")
        update_columns = [
            ("ID", 10),
            ("Content", 70),
            ("Created At", 22),
            ("Item IDs", 20),
        ]
        for col_idx, (title, width) in enumerate(update_columns, start=1):
            cell = updates_sheet.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            updates_sheet.column_dimensions[cell.column_letter].width = width

        for row_idx, update in enumerate(updates, start=2):
            updates_sheet.cell(row=row_idx, column=1, value=update.get("id"))
            updates_sheet.cell(row=row_idx, column=2, value=update.get("content"))
            updates_sheet.cell(row=row_idx, column=3, value=update.get("created_at"))
            updates_sheet.cell(row=row_idx, column=4, value=", ".join(map(str, update.get("item_ids", []))))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _group_items_by_status(items: List[Dict]) -> Dict[str, List[Dict]]:
    """Group exported items by status."""
    grouped: Dict[str, List[Dict]] = {}
    for item in items:
        grouped.setdefault(item.get("status_name") or "", []).append(item)
    return grouped


def _ordered_statuses(items_by_status: Dict[str, List[Dict]]) -> List[str]:
    """Return report statuses in the configured order."""
    ordered = [status for status in STATUS_COLUMN_ORDER if status in items_by_status]
    for status in sorted(items_by_status.keys()):
        if status not in STATUS_COLUMN_ORDER:
            ordered.append(status)
    return ordered


def _format_decision_history(decisions: List[Dict]) -> str:
    """Format decision history for a spreadsheet cell."""
    if not decisions:
        return "-"

    parts = []
    for decision in decisions:
        choice = decision.get("choice") or ""
        rationale = decision.get("rationale") or ""
        if rationale:
            parts.append(f"{choice}\n{rationale}")
        else:
            parts.append(choice)
    return "\n\n".join(part for part in parts if part)


def _format_status_card(item: Dict[str, Any]) -> str:
    """Format a compact item card for the status sheet."""
    title = item.get("title", "")
    priority = item.get("priority")
    parts = [f"#{item.get('id')} - {title}"]
    if priority is not None:
        parts.append(f"עדיפות: P{priority}")
    decisions = _format_decision_history(item.get("decisions", []))
    if decisions != "-":
        parts.append("היסטוריית החלטות:")
        parts.append(decisions)
    return "\n".join(parts)


def _format_item_age_days(created_at: Any) -> Optional[int]:
    """Return rounded age in days from created_at until now."""
    if not created_at:
        return None

    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at)
        except (ValueError, TypeError):
            return None

    if not isinstance(created_at, datetime):
        return None

    age_days = round((datetime.now() - created_at).total_seconds() / 86400)
    return max(0, int(age_days))


def get_mime_type(format: str) -> str:
    """Get MIME type for export format.

    Args:
        format: Export format

    Returns:
        MIME type string
    """
    format = format.lower()
    mime_types = {
        "json": "application/json",
        "yaml": "text/yaml",
        "markdown": "text/markdown",
        "md": "text/markdown",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    return mime_types.get(format, "text/plain")


def get_file_extension(format: str) -> str:
    """Get file extension for export format.

    Args:
        format: Export format

    Returns:
        File extension (including dot)
    """
    format = format.lower()
    extensions = {
        "json": ".json",
        "yaml": ".yaml",
        "markdown": ".md",
        "md": ".md",
        "xlsx": ".xlsx"
    }
    return extensions.get(format, ".txt")
